"""
Fonctions utilitaires pour écrire des fichiers Excel propres et lisibles,
avec une mise en forme professionnelle (largeur de colonnes, en-têtes stylés).
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def ecrire_excel_propre(dataframe, chemin_fichier):
    """
    Écrit un DataFrame dans un fichier Excel avec une mise en forme propre :
    - largeur de colonnes ajustée automatiquement au contenu
    - en-têtes en gras avec un fond coloré
    - alignement centré pour les en-têtes
    """
    dataframe.to_excel(chemin_fichier, index=False)

    classeur = load_workbook(chemin_fichier)
    feuille = classeur.active

    # Style des en-têtes (ligne 1) : fond bleu foncé, texte blanc et gras
    style_entete_police = Font(bold=True, color="FFFFFF", size=11)
    style_entete_fond = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    style_entete_alignement = Alignment(horizontal="center", vertical="center")

    for cellule in feuille[1]:  # feuille[1] = toute la première ligne
        cellule.font = style_entete_police
        cellule.fill = style_entete_fond
        cellule.alignment = style_entete_alignement

    # Largeur de colonnes ajustée au contenu le plus long
    for colonne in feuille.columns:
        longueur_max = max(len(str(cellule.value)) for cellule in colonne)
        lettre_colonne = colonne[0].column_letter
        feuille.column_dimensions[lettre_colonne].width = longueur_max + 4

    # Hauteur de la ligne d'en-tête un peu plus grande
    feuille.row_dimensions[1].height = 22

    # On fige la première ligne pour qu'elle reste visible en scrollant
    feuille.freeze_panes = "A2"

    classeur.save(chemin_fichier)